#!/usr/bin/env python3
"""
Claude Code 交互分析脚本
读取 merged.json 文件，生成结构化分析表格（Excel 和 Markdown 格式）
"""

import json
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font


def extract_user_text(message):
    """从 message 中提取 user 的纯文本内容"""
    if not message or not isinstance(message, list):
        return ""

    user_texts = []
    for msg in message:
        if msg.get("role") == "user":
            content = msg.get("content", "")
            if isinstance(content, str):
                user_texts.append(content)
            elif isinstance(content, list):
                for item in content:
                    if isinstance(item, dict) and item.get("type") == "text":
                        user_texts.append(item.get("text", ""))
                    elif isinstance(item, str):
                        user_texts.append(item)

    return "\n".join(user_texts)


def determine_role(message):
    """确定 message 中的主要角色"""
    if not message or not isinstance(message, list):
        return ""

    # 优先返回 assistant（因为包含 tool_use）
    for msg in message:
        if msg.get("role") == "assistant":
            return "assistant"

    # 否则返回 user
    for msg in message:
        if msg.get("role") == "user":
            return "user"

    return ""


def extract_tool_info(tool_use):
    """从 tool_use 中提取工具数量和名称"""
    if not tool_use or not isinstance(tool_use, list):
        return 0, ""

    tool_count = len(tool_use)
    tool_names = []

    for tool in tool_use:
        if isinstance(tool, dict):
            # 优先获取 subagent_type
            if "subagent_type" in tool:
                tool_names.append(tool["subagent_type"])
            # 否则获取第一个 key
            elif tool:
                first_key = next(iter(tool.keys()), "")
                if first_key:
                    tool_names.append(first_key)

    return tool_count, ", ".join(tool_names) if tool_names else ""


def process_data(data):
    """处理数据并生成表格行"""
    rows = []

    for step_name, interactions in data.items():
        for turn_index, interaction in enumerate(interactions):
            message = interaction.get("message", [])
            thinking = interaction.get("thinking", "")
            text = interaction.get("text", "")
            tool_use = interaction.get("tool_use", [])

            role = determine_role(message)
            user_text = extract_user_text(message)
            tool_count, tool_names = extract_tool_info(tool_use)
            raw_tool_json = json.dumps(tool_use, ensure_ascii=False) if tool_use else ""

            row = {
                "step": step_name,
                "turn_index": turn_index,
                "role": role,
                "user_text": user_text,
                "assistant_text": text,
                "thinking": thinking,
                "tool_count": tool_count,
                "tool_names": tool_names,
                "raw_tool_json": raw_tool_json
            }
            rows.append(row)

    return rows


def save_to_excel(df, filename):
    """保存到 Excel 文件，并设置格式"""
    df.to_excel(filename, index=False, engine='openpyxl')

    # 加载工作簿设置格式
    wb = load_workbook(filename)
    ws = wb.active

    # 首行加粗
    for cell in ws[1]:
        cell.font = Font(bold=True)

    # 自动设置列宽
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value:
                    # 计算字符长度，中文字符算2个宽度
                    cell_length = 0
                    for char in str(cell.value):
                        if '\u4e00' <= char <= '\u9fff':
                            cell_length += 2
                        else:
                            cell_length += 1
                    max_length = max(max_length, cell_length)
            except:
                pass
        # 设置列宽，最大50
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(filename)


def save_to_markdown(df, filename):
    """保存到 Markdown 文件"""
    with open(filename, 'w', encoding='utf-8') as f:
        f.write("# Claude Code 交互分析表\n\n")
        # 使用 tabulate 格式化表格
        markdown_table = df.to_markdown(index=False)
        f.write(markdown_table)
        f.write("\n")


def main():
    # 读取 JSON 文件
    input_file = "merged.json"
    with open(input_file, 'r', encoding='utf-8') as f:
        data = json.load(f)

    # 处理数据
    rows = process_data(data)

    # 创建 DataFrame
    columns = ["step", "turn_index", "role", "user_text", "assistant_text",
               "thinking", "tool_count", "tool_names", "raw_tool_json"]
    df = pd.DataFrame(rows, columns=columns)

    # 将空字符串保持为空字符串，避免显示为 NaN
    df = df.fillna("")

    # 保存到 Excel
    excel_file = "analysis.xlsx"
    save_to_excel(df, excel_file)
    print(f"Excel 文件已生成: {excel_file}")

    # 保存到 Markdown
    md_file = "analysis.md"
    save_to_markdown(df, md_file)
    print(f"Markdown 文件已生成: {md_file}")

    # 打印统计信息
    print(f"\n统计信息:")
    print(f"  总步骤数: {len(data)}")
    print(f"  总交互数: {len(df)}")
    print(f"  总工具调用数: {df['tool_count'].sum()}")


if __name__ == "__main__":
    main()
